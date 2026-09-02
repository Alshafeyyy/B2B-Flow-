import logging
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("Exporter")

def export_pipeline_to_excel(
    companies_data: List[Dict[str, Any]],
    contacts_briefs_data: List[Dict[str, Any]],
    output_filename: str = "ALX_Enterprise_Sourcing_Pipeline.xlsx",
    sheet1_name: str = "Company Qualification",
    sheet2_name: str = "Enriched Contacts & Briefs"
) -> str:
    """
    Exports 2-phase pipeline output into a styled multi-sheet Excel workbook.
    Sheet 1: Company Qualification & Role Suggestions (or a company dossier, for the Account Deep-Dive flow)
    Sheet 2: Enriched Contacts & Deep Pre-Call Briefs
    """
    logger.info(f"Exporting pipeline data to Excel workbook: {output_filename}...")

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # Sheet 1: Company Sourcing & Qualification
        if companies_data:
            df_comp = pd.DataFrame(companies_data)
            df_comp.to_excel(writer, sheet_name=sheet1_name, index=False)
        else:
            pd.DataFrame([{"Message": "No companies processed."}]).to_excel(writer, sheet_name=sheet1_name, index=False)

        # Sheet 2: Enriched Contacts & Briefs
        if contacts_briefs_data:
            df_contacts = pd.DataFrame(contacts_briefs_data)
            df_contacts.to_excel(writer, sheet_name=sheet2_name, index=False)
        else:
            pd.DataFrame([{"Message": "No contacts retrieved for 'Go' companies."}]).to_excel(writer, sheet_name=sheet2_name, index=False)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_filename)

        fill_go = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")     # Light green
        fill_nogo = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")   # Light red
        fill_review = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # Light yellow
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid") # Deep navy header
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            ws.views.sheetView[0].showGridLines = True

            # Format Header Row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Find status column for coloring
            status_col_idx = None
            for col in range(1, ws.max_column + 1):
                h_val = str(ws.cell(row=1, column=col).value or "").strip().lower()
                if "status" in h_val or "verdict" in h_val:
                    status_col_idx = col
                    break

            for row in range(2, ws.max_row + 1):
                if status_col_idx:
                    cell = ws.cell(row=row, column=status_col_idx)
                    val = str(cell.value or "").strip().lower()
                    if "go" in val and "no" not in val:
                        cell.fill = fill_go
                    elif "no" in val:
                        cell.fill = fill_nogo
                    elif "review" in val:
                        cell.fill = fill_review

                # Wrap text for long content cells
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

            # Adjust Column Widths
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                header_name = str(col[0].value or "")
                if any(k in header_name for k in ["Brief", "Angle", "Insights", "Position", "Roles", "Sources", "Organization", "Developments", "Points", "Presence", "Background", "Controversies", "Overview"]):
                    ws.column_dimensions[col_letter].width = 45
                elif "Reason" in header_name or "Description" in header_name:
                    ws.column_dimensions[col_letter].width = 35
                else:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 30)

        wb.save(output_filename)
        logger.info(f"Successfully formatted multi-sheet Excel file: {output_filename}")
    except Exception as e:
        logger.warning(f"Excel styling step warning: {e}")

    return output_filename
